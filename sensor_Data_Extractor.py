import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import os
import time
from datetime import datetime

class SensorDataAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("Sensor Data Extractor")
        self.root.geometry("1200x800")
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create first tab (original functionality)
        self.tab1 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab1, text="Single Sensor Analysis")
        
        # Create second tab (new batch functionality)
        self.tab2 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab2, text="Batch Extract from Template")
        
        # Create third tab (professional plotting)
        self.tab3 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab3, text="Professional Plotting")
        
        # Create fourth tab (tutorial)
        self.tab4 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab4, text="User Tutorial")
        
        # Create fifth tab (about us)
        self.tab5 = ttk.Frame(self.notebook)
        self.notebook.add(self.tab5, text="About")
        
        # Initialize variables for all tabs
        self.data = None
        self.current_window_data = None
        self.excel_writer = None
        self.excel_filename = None
        self.window_ranges = []
        self.current_sensor_index = 0
        
        # Setup all tabs
        self.setup_tab1()
        self.setup_tab2()
        self.setup_tab3()
        self.setup_tab4()  # New tutorial tab
        self.setup_tab5()  # New about tab
        
    def setup_tab1(self):
        # Main frame for tab1
        main_frame = ttk.Frame(self.tab1, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.tab1.columnconfigure(0, weight=1)
        self.tab1.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # File selection section
        ttk.Label(main_frame, text="Select Data File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.file_path = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.file_path, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(main_frame, text="Browse", command=self.browse_file).grid(row=0, column=2, pady=5, padx=5)
        
        # Plot controls section
        plot_frame = ttk.LabelFrame(main_frame, text="Plot Controls", padding="10")
        plot_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        plot_frame.columnconfigure(1, weight=1)
        
        ttk.Button(plot_frame, text="Plot All Data", command=self.plot_all_data).grid(row=0, column=0, pady=5, padx=5)
        
        ttk.Label(plot_frame, text="Window Range (Index):").grid(row=0, column=1, sticky=tk.W, pady=5, padx=5)
        self.start_range = ttk.Entry(plot_frame, width=10)
        self.start_range.grid(row=0, column=2, pady=5, padx=5)
        ttk.Label(plot_frame, text="to").grid(row=0, column=3, pady=5, padx=5)
        self.end_range = ttk.Entry(plot_frame, width=10)
        self.end_range.grid(row=0, column=4, pady=5, padx=5)
        
        ttk.Button(plot_frame, text="Plot Selected Window", command=self.plot_window).grid(row=0, column=5, pady=5, padx=5)
        ttk.Button(plot_frame, text="Add Window to List", command=self.add_window_to_list).grid(row=0, column=6, pady=5, padx=5)
        
        # Add info label for expression support
        info_label = ttk.Label(plot_frame, text="Supports expressions like: 400+600, 400-20", foreground="gray", font=("Arial", 8))
        info_label.grid(row=1, column=0, columnspan=7, pady=2)
        
        # Saved windows display
        self.windows_listbox = tk.Listbox(plot_frame, height=4, width=50)
        self.windows_listbox.grid(row=2, column=0, columnspan=6, sticky=(tk.W, tk.E), pady=5)
        ttk.Button(plot_frame, text="Remove Selected", command=self.remove_selected_window).grid(row=2, column=6, pady=5, padx=5)
        
        # Excel export section
        export_frame = ttk.LabelFrame(main_frame, text="Export to Excel", padding="10")
        export_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        export_frame.columnconfigure(1, weight=1)
        
        ttk.Label(export_frame, text="Sheet Name:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.sheet_name = ttk.Entry(export_frame, width=20)
        self.sheet_name.grid(row=0, column=1, sticky=tk.W, pady=5, padx=5)
        self.sheet_name.insert(0, "Window_1")
        
        ttk.Button(export_frame, text="Save Current Window to Excel", command=self.save_to_excel).grid(row=0, column=2, pady=5, padx=5)
        ttk.Button(export_frame, text="Save All Windows to Excel", command=self.save_all_windows_to_excel).grid(row=0, column=3, pady=5, padx=5)
        ttk.Button(export_frame, text="New Excel File", command=self.new_excel_file).grid(row=0, column=4, pady=5, padx=5)
        
        # Excel file status
        self.excel_status_var = tk.StringVar()
        self.excel_status_var.set("No Excel file created yet")
        excel_status_label = ttk.Label(export_frame, textvariable=self.excel_status_var, foreground="red")
        excel_status_label.grid(row=1, column=0, columnspan=5, sticky=tk.W, pady=5)
        
        # Status section
        self.status_var = tk.StringVar()
        self.status_var.set("Please load a data file to begin")
        status_label = ttk.Label(main_frame, textvariable=self.status_var, foreground="blue")
        status_label.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=5)
        
        # Plot area
        plot_area_frame = ttk.LabelFrame(main_frame, text="Data Visualization", padding="10")
        plot_area_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        plot_area_frame.columnconfigure(0, weight=1)
        plot_area_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=2)
        
        # Create figure and canvas with 2x height
        self.fig, self.ax = plt.subplots(figsize=(10, 12))
        self.canvas = FigureCanvasTkAgg(self.fig, master=plot_area_frame)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
    def setup_tab2(self):
        """Setup the batch extraction tab"""
        # Main frame for tab2
        main_frame = ttk.Frame(self.tab2, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.tab2.columnconfigure(0, weight=1)
        self.tab2.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Template file section
        template_frame = ttk.LabelFrame(main_frame, text="Template File", padding="10")
        template_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        template_frame.columnconfigure(1, weight=1)
        
        ttk.Label(template_frame, text="Select Template File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.template_file_path = tk.StringVar()
        ttk.Entry(template_frame, textvariable=self.template_file_path, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(template_frame, text="Browse Template", command=self.browse_template_file).grid(row=0, column=2, pady=5, padx=5)
        
        # Template type selection
        template_type_frame = ttk.Frame(template_frame)
        template_type_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        self.template_type = tk.StringVar(value="excel")
        ttk.Radiobutton(template_type_frame, text="Excel Template", variable=self.template_type, value="excel").grid(row=0, column=0, padx=10)
        ttk.Radiobutton(template_type_frame, text="Text Template", variable=self.template_type, value="text").grid(row=0, column=1, padx=10)
        
        ttk.Button(template_frame, text="Load Template Info", command=self.load_template_info).grid(row=2, column=0, columnspan=3, pady=10)
        
        # Template info display
        self.template_info_text = tk.Text(template_frame, height=8, width=80)
        self.template_info_text.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        scrollbar = ttk.Scrollbar(template_frame, orient="vertical", command=self.template_info_text.yview)
        scrollbar.grid(row=3, column=3, sticky=(tk.N, tk.S), pady=5)
        self.template_info_text.configure(yscrollcommand=scrollbar.set)
        
        # New sensor data section
        data_frame = ttk.LabelFrame(main_frame, text="New Sensor Data", padding="10")
        data_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        data_frame.columnconfigure(1, weight=1)
        
        ttk.Label(data_frame, text="Select New Sensor Data File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.new_data_file_path = tk.StringVar()
        ttk.Entry(data_frame, textvariable=self.new_data_file_path, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(data_frame, text="Browse Data", command=self.browse_new_data_file).grid(row=0, column=2, pady=5, padx=5)
        
        # Output file section
        output_frame = ttk.LabelFrame(main_frame, text="Output File", padding="10")
        output_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        output_frame.columnconfigure(1, weight=1)
        
        ttk.Label(output_frame, text="Output Excel File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.output_file_path = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.output_file_path, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(output_frame, text="Browse Output", command=self.browse_output_file).grid(row=0, column=2, pady=5, padx=5)
        
        # Process button
        process_frame = ttk.Frame(main_frame)
        process_frame.grid(row=3, column=0, columnspan=3, pady=20)
        
        ttk.Button(process_frame, text="EXTRACT DATA FROM TEMPLATE", 
                  command=self.extract_from_template, 
                  style="Accent.TButton").grid(row=0, column=0, pady=10)
        
        # Create a style for the accent button
        style = ttk.Style()
        style.configure("Accent.TButton", foreground="white", background="#0078D4")
        
        # Progress and timing section
        progress_frame = ttk.LabelFrame(main_frame, text="Progress & Timing", padding="10")
        progress_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        progress_frame.columnconfigure(0, weight=1)
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
        self.progress_bar.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # Timing labels
        timing_frame = ttk.Frame(progress_frame)
        timing_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.start_time_var = tk.StringVar(value="Start Time: --:--:--")
        self.end_time_var = tk.StringVar(value="End Time: --:--:--")
        self.duration_var = tk.StringVar(value="Duration: --")
        
        ttk.Label(timing_frame, textvariable=self.start_time_var).grid(row=0, column=0, sticky=tk.W, padx=5)
        ttk.Label(timing_frame, textvariable=self.end_time_var).grid(row=0, column=1, sticky=tk.W, padx=20)
        ttk.Label(timing_frame, textvariable=self.duration_var).grid(row=0, column=2, sticky=tk.W, padx=5)
        
        # Status section for tab2
        self.tab2_status_var = tk.StringVar()
        self.tab2_status_var.set("Please load a template file to begin")
        tab2_status_label = ttk.Label(main_frame, textvariable=self.tab2_status_var, foreground="blue")
        tab2_status_label.grid(row=5, column=0, columnspan=3, sticky=tk.W, pady=5)
        
        # Variables for template data
        self.template_sheets = []  # List of (sheet_name, start_index, end_index, data_points)
        self.new_sensor_data = None
        
    def setup_tab3(self):
        """Setup the professional plotting tab"""
        # Main frame for tab3
        main_frame = ttk.Frame(self.tab3, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.tab3.columnconfigure(0, weight=1)
        self.tab3.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Excel file selection section
        excel_frame = ttk.LabelFrame(main_frame, text="Excel File Selection", padding="10")
        excel_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        excel_frame.columnconfigure(1, weight=1)
        
        ttk.Label(excel_frame, text="Select Excel File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.plot_excel_path = tk.StringVar()
        ttk.Entry(excel_frame, textvariable=self.plot_excel_path, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(excel_frame, text="Browse Excel", command=self.browse_plot_excel).grid(row=0, column=2, pady=5, padx=5)
        
        ttk.Button(excel_frame, text="Load Excel Sheets", command=self.load_excel_sheets).grid(row=1, column=0, columnspan=3, pady=10)
        
        # Sheets selection section
        sheets_frame = ttk.LabelFrame(main_frame, text="Sheet Selection", padding="10")
        sheets_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        sheets_frame.columnconfigure(0, weight=1)
        
        # Listbox for sheets with scrollbar
        listbox_frame = ttk.Frame(sheets_frame)
        listbox_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        listbox_frame.columnconfigure(0, weight=1)
        listbox_frame.rowconfigure(0, weight=1)
        
        self.sheets_listbox = tk.Listbox(listbox_frame, selectmode=tk.MULTIPLE, height=8)
        self.sheets_listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        listbox_scrollbar = ttk.Scrollbar(listbox_frame, orient="vertical", command=self.sheets_listbox.yview)
        listbox_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.sheets_listbox.configure(yscrollcommand=listbox_scrollbar.set)
        
        # Sheet control buttons
        sheet_controls_frame = ttk.Frame(sheets_frame)
        sheet_controls_frame.grid(row=1, column=0, columnspan=2, pady=5)
        
        ttk.Button(sheet_controls_frame, text="Select All", command=self.select_all_sheets).grid(row=0, column=0, padx=5)
        ttk.Button(sheet_controls_frame, text="Clear Selection", command=self.clear_sheet_selection).grid(row=0, column=1, padx=5)
        ttk.Button(sheet_controls_frame, text="Preview Selected", command=self.preview_selected_sheet).grid(row=0, column=2, padx=5)
        
        # Plot options section
        plot_options_frame = ttk.LabelFrame(main_frame, text="Plot Options", padding="10")
        plot_options_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        plot_options_frame.columnconfigure(1, weight=1)
        
        # Figure size options
        ttk.Label(plot_options_frame, text="Figure Size:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.figure_size = tk.StringVar(value="12x8")
        size_combo = ttk.Combobox(plot_options_frame, textvariable=self.figure_size, width=15)
        size_combo['values'] = ('8x6', '10x7', '12x8', '14x10', '16x12')
        size_combo.grid(row=0, column=1, sticky=tk.W, pady=5, padx=5)
        
        # DPI options
        ttk.Label(plot_options_frame, text="DPI (Resolution):").grid(row=0, column=2, sticky=tk.W, pady=5, padx=(20,0))
        self.dpi_value = tk.StringVar(value="300")
        dpi_combo = ttk.Combobox(plot_options_frame, textvariable=self.dpi_value, width=10)
        dpi_combo['values'] = ('150', '200', '300', '400', '600')
        dpi_combo.grid(row=0, column=3, sticky=tk.W, pady=5, padx=5)
        
        # File format options
        ttk.Label(plot_options_frame, text="File Format:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.file_format = tk.StringVar(value="png")
        format_combo = ttk.Combobox(plot_options_frame, textvariable=self.file_format, width=15)
        format_combo['values'] = ('png', 'pdf', 'svg', 'jpg')
        format_combo.grid(row=1, column=1, sticky=tk.W, pady=5, padx=5)
        
        # Plot style options
        ttk.Label(plot_options_frame, text="Plot Style:").grid(row=1, column=2, sticky=tk.W, pady=5, padx=(20,0))
        self.plot_style = tk.StringVar(value="seaborn-v0_8-whitegrid")
        style_combo = ttk.Combobox(plot_options_frame, textvariable=self.plot_style, width=15)
        style_combo['values'] = ('seaborn-v0_8-whitegrid', 'seaborn-v0_8-darkgrid', 'seaborn-v0_8-poster', 
                               'seaborn-v0_8-talk', 'ggplot', 'classic')
        style_combo.grid(row=1, column=3, sticky=tk.W, pady=5, padx=5)
        
        # Custom axis labels
        ttk.Label(plot_options_frame, text="X-Axis Label:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.x_axis_label = tk.StringVar(value="Index")
        ttk.Entry(plot_options_frame, textvariable=self.x_axis_label, width=15).grid(row=2, column=1, sticky=tk.W, pady=5, padx=5)
        
        ttk.Label(plot_options_frame, text="Y-Axis Label:").grid(row=2, column=2, sticky=tk.W, pady=5, padx=(20,0))
        self.y_axis_label = tk.StringVar(value="Sensor Value")
        ttk.Entry(plot_options_frame, textvariable=self.y_axis_label, width=15).grid(row=2, column=3, sticky=tk.W, pady=5, padx=5)
        
        # Output folder section
        output_frame = ttk.LabelFrame(main_frame, text="Output Folder", padding="10")
        output_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        output_frame.columnconfigure(1, weight=1)
        
        ttk.Label(output_frame, text="Save Plots To:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.plot_output_folder = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.plot_output_folder, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        ttk.Button(output_frame, text="Browse Folder", command=self.browse_plot_folder).grid(row=0, column=2, pady=5, padx=5)
        
        # Plot buttons section
        plot_buttons_frame = ttk.Frame(main_frame)
        plot_buttons_frame.grid(row=4, column=0, columnspan=3, pady=20)
        
        ttk.Button(plot_buttons_frame, text="PLOT SELECTED SHEETS", 
                  command=self.plot_selected_sheets, 
                  style="Accent.TButton").grid(row=0, column=0, padx=10)
        
        ttk.Button(plot_buttons_frame, text="PLOT ALL SHEETS", 
                  command=self.plot_all_sheets, 
                  style="Accent.TButton").grid(row=0, column=1, padx=10)
        
        # Progress section for tab3
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        progress_frame.columnconfigure(0, weight=1)
        
        # Progress bar
        self.plot_progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.plot_progress_bar.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # Progress labels
        self.plot_progress_var = tk.StringVar(value="Ready to plot")
        plot_progress_label = ttk.Label(progress_frame, textvariable=self.plot_progress_var)
        plot_progress_label.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=5)
        
        # Status section for tab3
        self.tab3_status_var = tk.StringVar()
        self.tab3_status_var.set("Please load an Excel file to begin")
        tab3_status_label = ttk.Label(main_frame, textvariable=self.tab3_status_var, foreground="blue")
        tab3_status_label.grid(row=6, column=0, columnspan=3, sticky=tk.W, pady=5)
        
        # Variables for plotting
        self.excel_sheets = []
        self.current_excel_data = None
        
    def setup_tab4(self):
        """Setup the tutorial tab"""
        # Main frame for tab4
        main_frame = ttk.Frame(self.tab4, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.tab4.columnconfigure(0, weight=1)
        self.tab4.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=1)
        
        # Create a text widget with scrollbar for the tutorial
        tutorial_frame = ttk.Frame(main_frame)
        tutorial_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        tutorial_frame.columnconfigure(0, weight=1)
        tutorial_frame.rowconfigure(0, weight=1)
        
        # Text widget for tutorial content
        self.tutorial_text = tk.Text(tutorial_frame, wrap=tk.WORD, width=100, height=30, font=("Arial", 10))
        self.tutorial_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbar for text widget
        scrollbar = ttk.Scrollbar(tutorial_frame, orient="vertical", command=self.tutorial_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.tutorial_text.configure(yscrollcommand=scrollbar.set)
        
        # Add tutorial content
        tutorial_content = """
SENSOR DATA EXTRACTOR - COMPREHENSIVE USER GUIDE

================================================================================
OVERVIEW
================================================================================

Welcome to the Sensor Data Extractor! This professional software tool is designed 
for processing, analyzing, and visualizing sensor data with efficiency and precision. 
The application features four main tabs, each serving a specific purpose in the 
data analysis workflow.

================================================================================
TAB 1: SINGLE SENSOR ANALYSIS
================================================================================

PURPOSE:
- Load and analyze individual sensor data files
- Visualize complete datasets or specific data windows
- Extract and export selected data ranges to Excel

STEP-BY-STEP GUIDE:

1. LOADING DATA:
   - Click "Browse" to select your sensor data file (TXT or CSV format)
   - The file should contain semicolon-separated values with two columns: Index and Value
   - Once loaded, the status will show the data shape and available range

2. PLOTTING DATA:
   - "Plot All Data": Visualizes the entire dataset
   - "Plot Selected Window": Plots a specific range defined by start and end indices
   - Range Input: Supports mathematical expressions (e.g., 400+600, 500-50)

3. WINDOW MANAGEMENT:
   - "Add Window to List": Saves the current window range for batch processing
   - "Remove Selected": Deletes unwanted window ranges from the list
   - The list preserves frequently used ranges for efficient workflow

4. EXPORT TO EXCEL:
   - "New Excel File": Creates a new Excel workbook for data export
   - "Save Current Window to Excel": Exports the currently displayed window
   - "Save All Windows to Excel": Batch exports all saved window ranges
   - Sheet names can be customized for better organization

PRO TIPS:
- Use mathematical expressions for precise range selection
- Save frequently used window ranges to speed up repetitive analyses
- Always create a new Excel file before attempting to export data

================================================================================
TAB 2: BATCH EXTRACT FROM TEMPLATE
================================================================================

PURPOSE:
- Extract multiple data windows from new sensor files using predefined templates
- Support for both Excel and text-based templates
- Batch processing with progress tracking and timing information

STEP-BY-STEP GUIDE:

1. TEMPLATE SELECTION:
   - Choose between "Excel Template" or "Text Template"
   - Excel Template: Uses existing Excel files with multiple sheets as templates
   - Text Template: Uses simple text files with "start,end" format on each line

2. LOAD TEMPLATE INFO:
   - Click "Load Template Info" to analyze the template file
   - The system displays detailed information about each window/sheet found
   - Verify the ranges and data points before proceeding

3. SELECT NEW DATA:
   - Choose the new sensor data file to extract windows from
   - Ensure the format matches the expected semicolon-separated structure

4. SET OUTPUT FILE:
   - Specify the destination Excel file for extracted data
   - The system will create multiple sheets matching the template structure

5. EXECUTE EXTRACTION:
   - Click "EXTRACT DATA FROM TEMPLATE" to start the batch process
   - Monitor progress through the progress bar and timing information
   - The system provides detailed completion reports

PRO TIPS:
- Use Excel templates when you need to replicate complex multi-sheet structures
- Text templates are ideal for simple range-based extractions
- Check the template info display to ensure all windows are correctly identified

================================================================================
TAB 3: PROFESSIONAL PLOTTING
================================================================================

PURPOSE:
- Generate publication-quality plots from Excel data
- Batch plotting of multiple sheets with consistent formatting
- Customizable plot appearance and output settings

STEP-BY-STEP GUIDE:

1. LOAD EXCEL FILE:
   - Select an Excel file containing data sheets to plot
   - Click "Load Excel Sheets" to populate the sheet list

2. SHEET SELECTION:
   - Use "Select All" or manually select multiple sheets
   - "Preview Selected" allows quick verification before batch processing
   - "Clear Selection" resets the selection

3. PLOT CUSTOMIZATION:
   - Figure Size: Adjust plot dimensions for different publication requirements
   - DPI: Control image resolution (higher DPI for print publications)
   - File Format: Choose between PNG, PDF, SVG, or JPG formats
   - Plot Style: Select from various matplotlib styles for different aesthetics
   - X-Axis Label: Customize the x-axis label text
   - Y-Axis Label: Customize the y-axis label text

4. OUTPUT CONFIGURATION:
   - Specify the folder where plots will be saved
   - Each sheet generates a separate plot file with the sheet name

5. EXECUTE PLOTTING:
   - "PLOT SELECTED SHEETS": Processes only the chosen sheets
   - "PLOT ALL SHEETS": Processes every sheet in the Excel file
   - Monitor progress through the progress bar and status updates

PRO TIPS:
- Use higher DPI (300-600) for publication-quality figures
- PDF format is ideal for vector-based graphics in publications
- Preview sheets to ensure data formatting is correct before batch processing
- Custom axis labels help create professional-looking figures for specific applications

================================================================================
BEST PRACTICES & TROUBLESHOOTING
================================================================================

DATA FORMAT REQUIREMENTS:
- Sensor files: Semicolon-separated, two columns (Index, Value), no header
- Excel templates: Must contain 'Index' column in each sheet
- Text templates: Simple "start,end" format on each line

COMMON ISSUES:
- "No data found in range": Check that the range values exist in your data
- Excel export errors: Ensure you've created a new Excel file first
- Plotting failures: Verify that sheets contain 'Index' and 'Value' columns

PERFORMANCE OPTIMIZATION:
- For large datasets, use specific ranges rather than plotting all data
- Close other applications when processing very large files
- Use text templates for faster batch processing of simple ranges

================================================================================
SUPPORT
================================================================================

For technical support, bug reports, or feature requests, please contact:
Javad Amanabadi
Email: j.amanabadi@aut.ac.ir or j.amanabadi@gmail.com

This software was developed to support academic and industrial research in 
structural engineering and sensor data analysis.

Version: 1.0
Release Date: October 2025
"""
        
        self.tutorial_text.insert(1.0, tutorial_content)
        self.tutorial_text.config(state='disabled')  # Make read-only
        
    def setup_tab5(self):
        """Setup the about tab"""
        # Main frame for tab5
        main_frame = ttk.Frame(self.tab5, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.tab5.columnconfigure(0, weight=1)
        self.tab5.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Sensor Data Extractor", 
                               font=("Arial", 24, "bold"), foreground="#2C3E50")
        title_label.grid(row=0, column=0, pady=(0, 20))
        
        # Version and date
        version_label = ttk.Label(main_frame, text="Version 1.0 | October 2025", 
                                 font=("Arial", 12), foreground="#7F8C8D")
        version_label.grid(row=1, column=0, pady=(0, 30))
        
        # About frame
        about_frame = ttk.LabelFrame(main_frame, text="About the Developer", padding="15")
        about_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=10)
        about_frame.columnconfigure(0, weight=1)
        
        # Developer information
        dev_info = """
Javad Amanabadi
PhD in Structural Engineering
Amirkabir University of Technology (Tehran Polytechnic)

SPECIALIZATION:
• Structural Health Monitoring
• Sensor Data Analysis
• Signal Processing
• Finite Element Analysis
• Experimental Mechanics

CONTACT INFORMATION:
Email: j.amanabadi@aut.ac.ir
       j.amanabadi@gmail.com

AFFILIATION:
Department of Civil and Environmental Engineering
Amirkabir University of Technology
Tehran, Iran
"""
        info_label = ttk.Label(about_frame, text=dev_info, font=("Arial", 10), justify=tk.LEFT)
        info_label.grid(row=0, column=0, sticky=tk.W)
        
        # Software description frame
        software_frame = ttk.LabelFrame(main_frame, text="Software Information", padding="15")
        software_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=10)
        software_frame.columnconfigure(0, weight=1)
        
        software_info = """
Sensor Data Extractor

PURPOSE:
This professional software tool is designed for comprehensive analysis and 
visualization of sensor data in structural engineering and related fields. 
It streamlines the process of data extraction, window selection, and 
professional-quality plot generation.

KEY FEATURES:
• Single sensor data extractor and visualization
• Batch data extraction using templates
• Professional plotting for publications
• Support for multiple data formats
• Efficient Excel integration
• User-friendly interface

TECHNOLOGY STACK:
• Python 3.x
• Tkinter for GUI
• Pandas for data processing
• Matplotlib for visualization
• OpenPyXL for Excel integration

LICENSE:
This software is provided for academic and research purposes.
For commercial use, please contact the developer.

ACKNOWLEDGEMENT:
Developed to support research in structural health monitoring 
and sensor data analysis at Amirkabir University of Technology.
"""
        software_label = ttk.Label(software_frame, text=software_info, font=("Arial", 9), justify=tk.LEFT)
        software_label.grid(row=0, column=0, sticky=tk.W)
        
        # Copyright
        copyright_label = ttk.Label(main_frame, text="© 2025 Javad Amanabadi. All Rights Reserved.", 
                                   font=("Arial", 9), foreground="#7F8C8D")
        copyright_label.grid(row=4, column=0, pady=(20, 0))

    # ===== TAB 1 METHODS (ORIGINAL FUNCTIONALITY - SIMPLIFIED) =====
    
    def evaluate_expression(self, expression):
        """Evaluate mathematical expressions like 400+600 or 400-20"""
        try:
            # Remove any whitespace
            expression = expression.replace(' ', '')
            
            # Safe evaluation - only allow basic arithmetic operations
            allowed_chars = set('0123456789+-*/(). ')
            if not all(c in allowed_chars for c in expression):
                raise ValueError("Invalid characters in expression")
            
            # Evaluate the expression
            result = eval(expression)
            return int(result)
        except Exception as e:
            raise ValueError(f"Invalid expression: {expression}")
    
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Select Sensor Data File",
            filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
            self.load_data()
    
    def load_data(self):
        try:
            # Read semicolon separated data with no header
            self.data = pd.read_csv(self.file_path.get(), sep=';', header=None, names=['Index', 'Value'])
            
            # Convert to numeric types to ensure proper plotting
            self.data['Index'] = pd.to_numeric(self.data['Index'])
            self.data['Value'] = pd.to_numeric(self.data['Value'])
            
            self.status_var.set(f"Data loaded successfully! Shape: {self.data.shape}")
            
            # Update range suggestions
            if len(self.data) > 0:
                self.start_range.delete(0, tk.END)
                self.end_range.delete(0, tk.END)
                self.start_range.insert(0, str(self.data['Index'].iloc[0]))
                self.end_range.insert(0, str(self.data['Index'].iloc[-1]))
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load data: {str(e)}")
            self.status_var.set("Error loading data file")
    
    def plot_all_data(self):
        if self.data is None:
            messagebox.showwarning("Warning", "Please load data first!")
            return
            
        self.ax.clear()
        
        # Plot the data using Index as x-axis and Value as y-axis
        self.ax.plot(self.data['Index'], self.data['Value'], label='Sensor Data', color='blue', linewidth=1)
        
        self.ax.set_title("All Sensor Data")
        self.ax.set_xlabel("Index")
        self.ax.set_ylabel("Sensor Value")
        self.ax.legend()
        self.ax.grid(True, alpha=0.3)
        
        self.canvas.draw()
        self.status_var.set("All data plotted successfully")
    
    def plot_window(self):
        if self.data is None:
            messagebox.showwarning("Warning", "Please load data first!")
            return
            
        try:
            # Evaluate expressions in range boxes
            start_expression = self.start_range.get()
            end_expression = self.end_range.get()
            
            start = self.evaluate_expression(start_expression)
            end = self.evaluate_expression(end_expression)
            
            # Update the entry boxes with calculated values
            self.start_range.delete(0, tk.END)
            self.start_range.insert(0, str(start))
            self.end_range.delete(0, tk.END)
            self.end_range.insert(0, str(end))
            
            # Find the actual row indices for the given index values
            start_mask = self.data['Index'] >= start
            end_mask = self.data['Index'] <= end
            window_mask = start_mask & end_mask
            
            if not window_mask.any():
                messagebox.showerror("Error", "No data found in the specified range!")
                return
                
            # Extract window data
            self.current_window_data = self.data[window_mask].copy()
            
            # Plot the window
            self.ax.clear()
            
            self.ax.plot(self.current_window_data['Index'], self.current_window_data['Value'], 
                       label=f'Window {start}-{end}', color='red', linewidth=2)
            
            self.ax.set_title(f"Sensor Data Window ({start} to {end})")
            self.ax.set_xlabel("Index")
            self.ax.set_ylabel("Sensor Value")
            self.ax.legend()
            self.ax.grid(True, alpha=0.3)
            
            self.canvas.draw()
            self.status_var.set(f"Window {start}-{end} plotted successfully (found {len(self.current_window_data)} data points)")
            
        except ValueError as e:
            messagebox.showerror("Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Please enter valid numeric range values! {str(e)}")
    
    def add_window_to_list(self):
        """Add the current window range to the list for reuse"""
        try:
            start = int(self.start_range.get())
            end = int(self.end_range.get())
            
            window_range = (start, end)
            if window_range not in self.window_ranges:
                self.window_ranges.append(window_range)
                self.windows_listbox.insert(tk.END, f"{start} - {end}")
                self.status_var.set(f"Window {start}-{end} added to list")
            else:
                self.status_var.set("Window range already in list")
                
        except ValueError:
            messagebox.showerror("Error", "Please enter valid window range first!")
    
    def remove_selected_window(self):
        """Remove selected window from the list"""
        selection = self.windows_listbox.curselection()
        if selection:
            index = selection[0]
            self.windows_listbox.delete(index)
            self.window_ranges.pop(index)
            self.status_var.set("Window removed from list")
    
    def save_windows_template(self):
        """Save the current window ranges to a template file"""
        if not self.window_ranges:
            messagebox.showwarning("Warning", "No window ranges to save!")
            return
            
        filename = filedialog.asksaveasfilename(
            title="Save Windows Template",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                with open(filename, 'w') as f:
                    for start, end in self.window_ranges:
                        f.write(f"{start},{end}\n")
                self.status_var.set(f"Windows template saved: {len(self.window_ranges)} windows")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save template: {str(e)}")
    
    def save_all_windows_to_excel(self):
        """Save all windows in the list to Excel at once"""
        if not self.window_ranges:
            messagebox.showwarning("Warning", "No windows to save!")
            return
            
        if self.data is None:
            messagebox.showwarning("Warning", "Please load data first!")
            return
            
        if not self.excel_filename:
            messagebox.showwarning("Warning", "Please create an Excel file first!")
            return
            
        try:
            for i, (start, end) in enumerate(self.window_ranges):
                # Extract window data
                start_mask = self.data['Index'] >= start
                end_mask = self.data['Index'] <= end
                window_mask = start_mask & end_mask
                
                if window_mask.any():
                    window_data = self.data[window_mask].copy()
                    
                    # Save to Excel
                    sheet_name = f"Window_{i+1}"
                    
                    with pd.ExcelWriter(self.excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        window_data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            self.status_var.set(f"Saved {len(self.window_ranges)} windows to Excel")
            messagebox.showinfo("Success", f"All {len(self.window_ranges)} windows saved successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save all windows: {str(e)}")
    
    def new_excel_file(self):
        try:
            # Close existing writer if it exists
            if hasattr(self, 'excel_writer') and self.excel_writer is not None:
                try:
                    self.excel_writer.close()
                except:
                    pass
                
            self.excel_filename = filedialog.asksaveasfilename(
                title="Create New Excel File",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if self.excel_filename:
                # Create a new Excel file with openpyxl
                from openpyxl import Workbook
                wb = Workbook()
                wb.save(self.excel_filename)
                
                self.excel_status_var.set(f"Excel file ready: {os.path.basename(self.excel_filename)}")
                self.status_var.set(f"New Excel file created: {self.excel_filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create Excel file: {str(e)}")
    
    def save_to_excel(self):
        if self.current_window_data is None:
            messagebox.showwarning("Warning", "Please plot a window first!")
            return
            
        if not self.excel_filename:
            messagebox.showwarning("Warning", "Please create an new Excel file first!")
            return
            
        try:
            sheet_name = self.sheet_name.get().strip()
            if not sheet_name:
                messagebox.showerror("Error", "Please enter a sheet name!")
                return
            
            # Check if file exists
            if not os.path.exists(self.excel_filename):
                messagebox.showerror("Error", "Excel file not found! Please create a new Excel file.")
                return
            
            # Check if sheet already exists
            from openpyxl import load_workbook
            try:
                wb = load_workbook(self.excel_filename)
                if sheet_name in wb.sheetnames:
                    response = messagebox.askyesno(
                        "Sheet Exists", 
                        f"Sheet '{sheet_name}' already exists. Do you want to overwrite it?"
                    )
                    if not response:
                        wb.close()
                        return
                    # Remove existing sheet
                    del wb[sheet_name]
            except Exception as e:
                messagebox.showerror("Error", f"Failed to check existing sheets: {str(e)}")
                return
            
            # Save the data to Excel
            try:
                # Use mode='a' to append to existing file
                with pd.ExcelWriter(self.excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    self.current_window_data.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:
                # If the above fails, try the manual method
                try:
                    wb = load_workbook(self.excel_filename)
                    if sheet_name in wb.sheetnames:
                        del wb[sheet_name]
                    wb.create_sheet(sheet_name)
                    ws = wb[sheet_name]
                    
                    # Write headers
                    ws.cell(row=1, column=1, value="Index")
                    ws.cell(row=1, column=2, value="Value")
                    
                    # Write data
                    for i, (idx, row) in enumerate(self.current_window_data.iterrows(), start=2):
                        ws.cell(row=i, column=1, value=row['Index'])
                        ws.cell(row=i, column=2, value=row['Value'])
                    
                    wb.save(self.excel_filename)
                    wb.close()
                except Exception as e2:
                    messagebox.showerror("Error", f"Failed to save to Excel: {str(e2)}")
                    return
            
            # Update sheet name for next window
            if '_' in self.sheet_name.get():
                try:
                    current_sheet_num = int(self.sheet_name.get().split('_')[-1])
                    next_sheet_num = current_sheet_num + 1
                    self.sheet_name.delete(0, tk.END)
                    self.sheet_name.insert(0, f"Window_{next_sheet_num}")
                except ValueError:
                    # If the pattern doesn't match, just keep the current name
                    pass
            
            self.status_var.set(f"Window saved to sheet '{sheet_name}' in {os.path.basename(self.excel_filename)}")
            messagebox.showinfo("Success", f"Data successfully saved to sheet '{sheet_name}'!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Excel: {str(e)}")

    # ===== TAB 2 METHODS (BATCH EXTRACTION) =====
    
    def browse_template_file(self):
        """Browse for template file (Excel or Text)"""
        if self.template_type.get() == "excel":
            filetypes = [("Excel files", "*.xlsx"), ("All files", "*.*")]
        else:
            filetypes = [("Text files", "*.txt"), ("All files", "*.*")]
            
        filename = filedialog.askopenfilename(
            title="Select Template File",
            filetypes=filetypes
        )
        if filename:
            self.template_file_path.set(filename)
            self.tab2_status_var.set("Template file loaded. Click 'Load Template Info' to analyze.")
    
    def browse_new_data_file(self):
        """Browse for new sensor data file"""
        filename = filedialog.askopenfilename(
            title="Select New Sensor Data File",
            filetypes=[("Text files", "*.txt"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.new_data_file_path.set(filename)
            self.tab2_status_var.set("New data file loaded.")
    
    def browse_output_file(self):
        """Browse for output Excel file"""
        filename = filedialog.asksaveasfilename(
            title="Select Output Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_file_path.set(filename)
            self.tab2_status_var.set("Output file selected.")
    
    def load_template_info(self):
        """Load and analyze template file (Excel or Text)"""
        if not self.template_file_path.get():
            messagebox.showerror("Error", "Please select a template file first!")
            return
        
        try:
            # Clear previous template data
            self.template_sheets = []
            template_file = self.template_file_path.get()
            
            # Build info text
            info_text = "Template Analysis Results:\n\n"
            
            if self.template_type.get() == "excel":
                info_text = self._load_excel_template(template_file, info_text)
            else:
                info_text = self._load_text_template(template_file, info_text)
            
            # Display the template info
            self.template_info_text.config(state='normal')
            self.template_info_text.delete(1.0, tk.END)
            self.template_info_text.insert(1.0, info_text)
            self.template_info_text.config(state='disabled')
            
            self.tab2_status_var.set(f"Template loaded: {len(self.template_sheets)} sheets/windows analyzed")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load template: {str(e)}")
    
    def _load_excel_template(self, template_file, info_text):
        """Load template from Excel file"""
        try:
            # Read the Excel file
            excel_file = pd.ExcelFile(template_file)
            sheet_names = excel_file.sheet_names
            
            info_text += f"Excel File: {os.path.basename(template_file)}\n"
            info_text += f"Total Sheets: {len(sheet_names)}\n\n"
            
            valid_sheets = 0
            for sheet_name in sheet_names:
                # Skip default empty sheets
                if sheet_name == 'Sheet' and len(sheet_names) > 1:
                    continue
                    
                try:
                    # Read sheet data
                    sheet_data = pd.read_excel(template_file, sheet_name=sheet_name)
                    
                    # Check if sheet has the required columns
                    if 'Index' in sheet_data.columns and not sheet_data.empty:
                        start_index = int(sheet_data['Index'].min())
                        end_index = int(sheet_data['Index'].max())
                        data_points = len(sheet_data)
                        
                        self.template_sheets.append((sheet_name, start_index, end_index, data_points))
                        valid_sheets += 1
                        
                        info_text += f"? Sheet: {sheet_name}\n"
                        info_text += f"  Range: {start_index} to {end_index}\n"
                        info_text += f"  Data Points: {data_points}\n"
                        info_text += "-" * 50 + "\n"
                    else:
                        info_text += f"? Sheet: {sheet_name} - No valid data (missing 'Index' column or empty)\n"
                        info_text += "-" * 50 + "\n"
                        
                except Exception as e:
                    info_text += f"? Sheet: {sheet_name} - Error reading: {str(e)}\n"
                    info_text += "-" * 50 + "\n"
            
            if valid_sheets == 0:
                info_text += "\n? No valid sheets found in the Excel file!\n"
                info_text += "Please make sure the Excel file contains sheets with 'Index' column and data.\n"
            else:
                info_text += f"\n? Successfully loaded {valid_sheets} sheets from template.\n"
                
        except Exception as e:
            info_text += f"\n? Error opening Excel file: {str(e)}\n"
            
        return info_text
    
    def _load_text_template(self, template_file, info_text):
        """Load template from Text file"""
        try:
            with open(template_file, 'r') as f:
                lines = f.readlines()
            
            info_text += f"Text File: {os.path.basename(template_file)}\n"
            info_text += f"Total Lines: {len(lines)}\n\n"
            
            valid_windows = 0
            for i, line in enumerate(lines):
                line = line.strip()
                if not line:  # Skip empty lines
                    continue
                    
                try:
                    # Parse "start,end" format
                    if ',' in line:
                        parts = line.split(',')
                        if len(parts) >= 2:
                            start_index = int(parts[0].strip())
                            end_index = int(parts[1].strip())
                            
                            sheet_name = f"Window_{i+1}"
                            self.template_sheets.append((sheet_name, start_index, end_index, "Will be extracted"))
                            valid_windows += 1
                            
                            info_text += f"? Window: {sheet_name}\n"
                            info_text += f"  Range: {start_index} to {end_index}\n"
                            info_text += f"  Data Points: Will be extracted from new data\n"
                            info_text += "-" * 50 + "\n"
                        else:
                            info_text += f"? Line {i+1}: Invalid format - {line}\n"
                            info_text += "-" * 50 + "\n"
                    else:
                        info_text += f"? Line {i+1}: Missing comma - {line}\n"
                        info_text += "-" * 50 + "\n"
                        
                except ValueError as e:
                    info_text += f"? Line {i+1}: Invalid numbers - {line}\n"
                    info_text += "-" * 50 + "\n"
            
            if valid_windows == 0:
                info_text += "\n? No valid window ranges found in the text file!\n"
                info_text += "Please make sure the file contains lines with format: 'start,end'\n"
            else:
                info_text += f"\n? Successfully loaded {valid_windows} windows from template.\n"
                
        except Exception as e:
            info_text += f"\n? Error reading text file: {str(e)}\n"
            
        return info_text
    
    def extract_from_template(self):
        """Extract data from new sensor data using template ranges"""
        # Validate inputs
        if not self.template_file_path.get():
            messagebox.showerror("Error", "Please select a template file first!")
            return
        
        if not self.template_sheets:
            messagebox.showerror("Error", "Please load template info first!")
            return
            
        if not self.new_data_file_path.get():
            messagebox.showerror("Error", "Please select a new sensor data file!")
            return
            
        if not self.output_file_path.get():
            messagebox.showerror("Error", "Please select an output Excel file!")
            return
        
        try:
            # Start progress bar and timing
            self.progress_bar.start()
            start_time = time.time()
            self.start_time_var.set(f"Start Time: {datetime.now().strftime('%H:%M:%S')}")
            self.end_time_var.set("End Time: --:--:--")
            self.duration_var.set("Duration: --")
            self.tab2_status_var.set("Starting extraction process...")
            self.root.update()
            
            # Load new sensor data
            self.tab2_status_var.set("Loading new sensor data...")
            self.new_sensor_data = pd.read_csv(self.new_data_file_path.get(), sep=';', header=None, names=['Index', 'Value'])
            self.new_sensor_data['Index'] = pd.to_numeric(self.new_sensor_data['Index'])
            self.new_sensor_data['Value'] = pd.to_numeric(self.new_sensor_data['Value'])
            
            # Create output Excel file
            from openpyxl import Workbook
            wb = Workbook()
            # Remove default sheet if it exists
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            wb.save(self.output_file_path.get())
            
            # Process each sheet from template
            successful_extractions = 0
            failed_extractions = 0
            total_sheets = len(self.template_sheets)
            
            for i, (sheet_name, start_index, end_index, _) in enumerate(self.template_sheets):
                try:
                    # Update progress
                    progress_percent = int((i + 1) / total_sheets * 100)
                    self.tab2_status_var.set(f"Processing {i+1}/{total_sheets} ({progress_percent}%): {sheet_name}...")
                    self.root.update()
                    
                    # Extract window data from new sensor data
                    start_mask = self.new_sensor_data['Index'] >= start_index
                    end_mask = self.new_sensor_data['Index'] <= end_index
                    window_mask = start_mask & end_mask
                    
                    if window_mask.any():
                        window_data = self.new_sensor_data[window_mask].copy()
                        
                        # Save to output Excel file
                        with pd.ExcelWriter(self.output_file_path.get(), engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            window_data.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        successful_extractions += 1
                    else:
                        failed_extractions += 1
                        self.tab2_status_var.set(f"Warning: No data found for range {start_index}-{end_index} in sheet {sheet_name}")
                        
                except Exception as e:
                    failed_extractions += 1
                    self.tab2_status_var.set(f"Error processing sheet {sheet_name}: {str(e)}")
            
            # Stop progress bar and update timing
            end_time = time.time()
            duration = end_time - start_time
            self.progress_bar.stop()
            self.end_time_var.set(f"End Time: {datetime.now().strftime('%H:%M:%S')}")
            
            # Format duration nicely
            if duration < 60:
                self.duration_var.set(f"Duration: {duration:.2f} seconds")
            else:
                minutes = int(duration // 60)
                seconds = duration % 60
                self.duration_var.set(f"Duration: {minutes}m {seconds:.2f}s")
            
            # Show completion message
            self.tab2_status_var.set(f"Extraction complete: {successful_extractions} successful, {failed_extractions} failed")
            messagebox.showinfo("Success", 
                              f"Data extraction completed!\n\n"
                              f"Successful extractions: {successful_extractions}\n"
                              f"Failed extractions: {failed_extractions}\n"
                              f"Time taken: {duration:.2f} seconds\n\n"
                              f"Output saved to: {self.output_file_path.get()}")
            
        except Exception as e:
            # Stop progress bar on error
            self.progress_bar.stop()
            messagebox.showerror("Error", f"Failed to extract data: {str(e)}")
            self.tab2_status_var.set("Extraction failed")

    # ===== TAB 3 METHODS (PROFESSIONAL PLOTTING) =====
    
    def browse_plot_excel(self):
        """Browse for Excel file for plotting"""
        filename = filedialog.askopenfilename(
            title="Select Excel File for Plotting",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.plot_excel_path.set(filename)
            self.tab3_status_var.set("Excel file loaded. Click 'Load Excel Sheets' to continue.")
    
    def browse_plot_folder(self):
        """Browse for output folder for plots"""
        folder = filedialog.askdirectory(title="Select Folder to Save Plots")
        if folder:
            self.plot_output_folder.set(folder)
            self.tab3_status_var.set(f"Output folder set: {folder}")
    
    def load_excel_sheets(self):
        """Load and display sheets from Excel file"""
        if not self.plot_excel_path.get():
            messagebox.showerror("Error", "Please select an Excel file first!")
            return
        
        try:
            # Read Excel file
            excel_file = pd.ExcelFile(self.plot_excel_path.get())
            self.excel_sheets = [sheet for sheet in excel_file.sheet_names if sheet != 'Sheet']
            
            # Update listbox
            self.sheets_listbox.delete(0, tk.END)
            for sheet in self.excel_sheets:
                self.sheets_listbox.insert(tk.END, sheet)
            
            self.tab3_status_var.set(f"Loaded {len(self.excel_sheets)} sheets from Excel file")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")
    
    def select_all_sheets(self):
        """Select all sheets in the listbox"""
        self.sheets_listbox.select_set(0, tk.END)
    
    def clear_sheet_selection(self):
        """Clear all selections in the listbox"""
        self.sheets_listbox.selection_clear(0, tk.END)
    
    def preview_selected_sheet(self):
        """Preview selected sheet in a new window"""
        selected_indices = self.sheets_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Warning", "Please select one sheet to preview!")
            return
        
        if len(selected_indices) > 1:
            messagebox.showwarning("Warning", "Please select only one sheet for preview!")
            return
        
        sheet_name = self.excel_sheets[selected_indices[0]]
        self._preview_sheet(sheet_name)
    
    def _preview_sheet(self, sheet_name):
        """Preview a single sheet in a new window"""
        try:
            # Read sheet data
            sheet_data = pd.read_excel(self.plot_excel_path.get(), sheet_name=sheet_name)
            
            if 'Index' not in sheet_data.columns or 'Value' not in sheet_data.columns:
                messagebox.showerror("Error", f"Sheet '{sheet_name}' does not contain required 'Index' and 'Value' columns!")
                return
            
            # Create preview window
            preview_window = tk.Toplevel(self.root)
            preview_window.title(f"Preview: {sheet_name}")
            preview_window.geometry("800x600")
            preview_window.transient(self.root)
            preview_window.grab_set()
            
            # Create figure for preview
            fig, ax = plt.subplots(figsize=(10, 6))
            
            # Plot data with integer formatting
            ax.plot(sheet_data['Index'], sheet_data['Value'], 
                   linewidth=2, color='#2E86AB', alpha=0.8)
            
            # Professional styling with integer formatting
            ax.set_xlabel(self.x_axis_label.get(), fontsize=12, fontweight='bold')
            ax.set_ylabel(self.y_axis_label.get(), fontsize=12, fontweight='bold')
            ax.set_title(f'Sensor Data - {sheet_name}', fontsize=14, fontweight='bold', pad=20)
            
            # Set integer formatting for both axes
            ax.ticklabel_format(style='plain', axis='both')
            ax.xaxis.set_major_formatter(plt.FormatStrFormatter('%d'))
            ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%d'))
            
            # Grid and aesthetics
            ax.grid(True, alpha=0.3, linestyle='--')
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            
            # Tight layout
            plt.tight_layout()
            
            # Embed plot in preview window
            canvas = FigureCanvasTkAgg(fig, master=preview_window)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Add close button
            close_button = ttk.Button(preview_window, text="Close Preview", 
                                    command=preview_window.destroy)
            close_button.pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to preview sheet: {str(e)}")
    
    def plot_selected_sheets(self):
        """Plot selected sheets"""
        selected_indices = self.sheets_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Warning", "Please select at least one sheet to plot!")
            return
        
        if not self.plot_output_folder.get():
            messagebox.showwarning("Warning", "Please select an output folder!")
            return
        
        selected_sheets = [self.excel_sheets[i] for i in selected_indices]
        self._plot_sheets(selected_sheets)
    
    def plot_all_sheets(self):
        """Plot all sheets"""
        if not self.excel_sheets:
            messagebox.showwarning("Warning", "No sheets available to plot!")
            return
        
        if not self.plot_output_folder.get():
            messagebox.showwarning("Warning", "Please select an output folder!")
            return
        
        self._plot_sheets(self.excel_sheets)
    
    def _plot_sheets(self, sheets_to_plot):
        """Internal method to plot sheets with professional styling"""
        try:
            # Set up progress bar
            total_sheets = len(sheets_to_plot)
            self.plot_progress_bar['maximum'] = total_sheets
            self.plot_progress_bar['value'] = 0
            
            # Parse figure size
            width, height = map(int, self.figure_size.get().split('x'))
            dpi = int(self.dpi_value.get())
            file_format = self.file_format.get()
            
            # Apply plot style
            plt.style.use(self.plot_style.get())
            
            successful_plots = 0
            failed_plots = 0
            
            for i, sheet_name in enumerate(sheets_to_plot):
                try:
                    # Update progress
                    self.plot_progress_bar['value'] = i + 1
                    self.plot_progress_var.set(f"Plotting {i+1}/{total_sheets}: {sheet_name}")
                    self.tab3_status_var.set(f"Creating plot for: {sheet_name}")
                    self.root.update()
                    
                    # Read sheet data
                    sheet_data = pd.read_excel(self.plot_excel_path.get(), sheet_name=sheet_name)
                    
                    # Create professional plot
                    fig, ax = plt.subplots(figsize=(width, height), dpi=dpi)
                    
                    # Plot data with integer formatting
                    if 'Index' in sheet_data.columns and 'Value' in sheet_data.columns:
                        ax.plot(sheet_data['Index'], sheet_data['Value'], 
                               linewidth=2, color='#2E86AB', alpha=0.8)
                        
                        # Professional styling with integer formatting
                        ax.set_xlabel(self.x_axis_label.get(), fontsize=12, fontweight='bold')
                        ax.set_ylabel(self.y_axis_label.get(), fontsize=12, fontweight='bold')
                        ax.set_title(f'Sensor Data - {sheet_name}', fontsize=14, fontweight='bold', pad=20)
                        
                        # Set integer formatting for both axes - NO scientific notation
                        ax.ticklabel_format(style='plain', axis='both')
                        ax.xaxis.set_major_formatter(plt.FormatStrFormatter('%d'))
                        ax.yaxis.set_major_formatter(plt.FormatStrFormatter('%d'))
                        
                        # Grid and aesthetics
                        ax.grid(True, alpha=0.3, linestyle='--')
                        ax.spines['top'].set_visible(False)
                        ax.spines['right'].set_visible(False)
                        
                        # Tight layout
                        plt.tight_layout()
                        
                        # Save plot
                        output_path = os.path.join(self.plot_output_folder.get(), f"{sheet_name}.{file_format}")
                        plt.savefig(output_path, dpi=dpi, bbox_inches='tight', 
                                  facecolor='white', edgecolor='none')
                        plt.close(fig)
                        
                        successful_plots += 1
                    else:
                        failed_plots += 1
                        self.tab3_status_var.set(f"Missing columns in sheet: {sheet_name}")
                        
                except Exception as e:
                    failed_plots += 1
                    self.tab3_status_var.set(f"Error plotting {sheet_name}: {str(e)}")
            
            # Completion message
            self.plot_progress_var.set(f"Completed: {successful_plots} successful, {failed_plots} failed")
            self.tab3_status_var.set(f"Plotting finished: {successful_plots} plots saved")
            
            messagebox.showinfo("Success", 
                              f"Plotting completed!\n\n"
                              f"Successful plots: {successful_plots}\n"
                              f"Failed plots: {failed_plots}\n\n"
                              f"Plots saved to: {self.plot_output_folder.get()}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create plots: {str(e)}")
            self.tab3_status_var.set("Plotting failed")

def main():
    root = tk.Tk()
    app = SensorDataAnalyzer(root)
    root.mainloop()

if __name__ == "__main__":
    main()